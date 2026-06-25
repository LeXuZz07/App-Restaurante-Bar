import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from datetime import datetime
import os
import database as db
import matplotlib
matplotlib.use('Agg') # Esto evita errores de interfaz gráfica en segundo plano
import matplotlib.pyplot as plt

def generar_excel_cierre(ventas, total, efectivo, tarjeta, tablet_id, productos_vendidos):
    ruta_db = db.get_db_path()
    base_path = os.path.dirname(ruta_db)
    nombre_subcarpeta = os.path.join(base_path, "Reportes_Cierre")
    os.makedirs(nombre_subcarpeta, exist_ok=True)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corte de Caja"
    
    # Datos generales
    ws["A1"] = f"CORTE DE CAJA - TABLET {tablet_id}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A4"], ws["B4"] = "TOTAL GENERAL:", total
    ws["A5"], ws["B5"] = "EFECTIVO:", efectivo
    ws["A6"], ws["B6"] = "TARJETA:", tarjeta
    
    # Encabezados en fila 8
    headers = ["Mesa", "Hora", "Método", "Total", "Detalle"]
    for col, text in enumerate(headers, 1):
        ws.cell(row=8, column=col, value=text).font = Font(bold=True)
    
    # Datos empezando en fila 9 (FORZAMOS POSICIÓN EXACTA)
    for row_idx, v in enumerate(ventas, 9):
        ws.cell(row=row_idx, column=1, value=f"Mesa {v[0]}")
        ws.cell(row=row_idx, column=2, value=v[3])
        ws.cell(row=row_idx, column=3, value=v[4])
        ws.cell(row=row_idx, column=4, value=v[2])
        ws.cell(row=row_idx, column=5, value=v[1])
        # Ajuste de celda para que no se corte
        ws.cell(row=row_idx, column=5).alignment = Alignment(wrap_text=True, vertical="top")
    
    ws.column_dimensions['E'].width = 50 
    
    # Hoja de Gráficas
    ws_g = wb.create_sheet("Estadísticas")
    ws_g["A1"], ws_g["B1"] = "Método", "Monto"
    ws_g["A2"], ws_g["B2"] = "Efectivo", efectivo
    ws_g["A3"], ws_g["B3"] = "Tarjeta", tarjeta
    
    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "Ingresos por Método"
    chart1.add_data(Reference(ws_g, min_col=2, min_row=1, max_row=3), titles_from_data=True)
    chart1.set_categories(Reference(ws_g, min_col=1, min_row=2, max_row=3))
    ws_g.add_chart(chart1, "D2")
    
    ws_g["D1"], ws_g["E1"] = "Producto", "Cantidad"
    for i, (prod, cant) in enumerate(productos_vendidos.items(), 2):
        ws_g.cell(row=i, column=4, value=prod)
        ws_g.cell(row=i, column=5, value=cant)
        
    chart2 = BarChart()
    chart2.type = "col"
    chart2.title = "Productos más Vendidos"
    data = Reference(ws_g, min_col=5, min_row=1, max_row=len(productos_vendidos)+1)
    cats = Reference(ws_g, min_col=4, min_row=2, max_row=len(productos_vendidos)+1)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    ws_g.add_chart(chart2, "D18")
    
    nombre_archivo = f"Corte_T{tablet_id}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    ruta_completa = os.path.join(nombre_subcarpeta, nombre_archivo)
    wb.save(ruta_completa)
    return ruta_completa

def leer_excel(ruta_completa):
    try:
        wb = openpyxl.load_workbook(ruta_completa, data_only=True)
        ws = wb.active
        filas = []
        for row in ws.iter_rows(values_only=True):
            filas.append([str(celda) if celda is not None else "" for celda in row])
        return filas
    except Exception as e:
        return None
    
def generar_graficas_imagenes(total_efe, total_tar, productos_vendidos, ruta_base):
    # 1. Gráfica de ingresos
    plt.figure(figsize=(6, 4))
    plt.bar(["Efectivo", "Tarjeta"], [total_efe, total_tar], color=['#28a745', '#007bff'])
    plt.title("Ingresos por Método")
    plt.ylabel("Monto ($)")
    ruta_ingresos = os.path.join(ruta_base, "grafica_ingresos.png")
    plt.savefig(ruta_ingresos)
    plt.close()
    
    # 2. Gráfica de productos (si hay ventas)
    ruta_productos = None
    if productos_vendidos:
        plt.figure(figsize=(7, 5))
        nombres = list(productos_vendidos.keys())
        cantidades = list(productos_vendidos.values())
        plt.bar(nombres, cantidades, color='#fd7e14')
        plt.title("Productos más vendidos")
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        ruta_productos = os.path.join(ruta_base, "grafica_productos.png")
        plt.savefig(ruta_productos)
        plt.close()
        
    return ruta_ingresos, ruta_productos