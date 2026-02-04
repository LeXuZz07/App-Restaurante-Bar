import flet as ft
import sqlite3
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
import os

# --- CONFIGURACIÓN ---
TABLET_ID = "01"

# --- CAPA DE DATOS (SQLite) ---
def init_db():
    conn = sqlite3.connect("pos_restaurante.db")
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS ventas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mesa_id INTEGER,
            detalle TEXT,
            total REAL,
            fecha TEXT,
            metodo_pago TEXT,
            cerrada INTEGER DEFAULT 0
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS items_activos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            mesa_id INTEGER,
            nombre TEXT,
            precio REAL,
            cantidad INTEGER,
            destino TEXT,
            enviado INTEGER
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS productos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT,
            precio REAL,
            categoria TEXT,
            destino TEXT
        )
    """)
    
    cursor.execute("SELECT count(*) FROM productos")
    if cursor.fetchone()[0] == 0:
        menu_inicial = [
            ("Cerveza", 55, "BEBIDAS", "BARRA"),
            ("Refresco", 35, "BEBIDAS", "BARRA"),
            ("Hamburguesa", 150, "COMIDA", "COCINA"),
            ("Tacos", 90, "COMIDA", "COCINA"),
            ("Pizza", 200, "COMIDA", "COCINA"),
            ("Pastel", 60, "POSTRES", "OTROS")
        ]
        cursor.executemany("INSERT INTO productos (nombre, precio, categoria, destino) VALUES (?,?,?,?)", menu_inicial)
    
    conn.commit()
    conn.close()

# --- FUNCIONES DE PERSISTENCIA ---
def db_guardar_item_activo(mesa, item):
    conn = sqlite3.connect("pos_restaurante.db")
    cursor = conn.cursor()
    cursor.execute("SELECT id, cantidad FROM items_activos WHERE mesa_id=? AND nombre=? AND enviado=0", (mesa, item['n']))
    res = cursor.fetchone()
    if res:
        cursor.execute("UPDATE items_activos SET cantidad=? WHERE id=?", (item['q'], res[0]))
    else:
        cursor.execute("INSERT INTO items_activos (mesa_id, nombre, precio, cantidad, destino, enviado) VALUES (?,?,?,?,?,?)",
                       (mesa, item['n'], item['p'], item['q'], item['d'], 0))
    conn.commit(); conn.close()

def db_remover_item_activo(mesa, nombre):
    conn = sqlite3.connect("pos_restaurante.db")
    cursor = conn.cursor()
    cursor.execute("SELECT id, cantidad FROM items_activos WHERE mesa_id=? AND nombre=? AND enviado=0", (mesa, nombre))
    res = cursor.fetchone()
    if res:
        if res[1] > 1: cursor.execute("UPDATE items_activos SET cantidad=? WHERE id=?", (res[1]-1, res[0]))
        else: cursor.execute("DELETE FROM items_activos WHERE id=?", (res[0],))
    conn.commit(); conn.close()

def db_marcar_enviados(mesa):
    conn = sqlite3.connect("pos_restaurante.db")
    conn.cursor().execute("UPDATE items_activos SET enviado=1 WHERE mesa_id=?", (mesa,))
    conn.commit(); conn.close()

def db_limpiar_mesa(mesa):
    conn = sqlite3.connect("pos_restaurante.db")
    conn.cursor().execute("DELETE FROM items_activos WHERE mesa_id=?", (mesa,))
    conn.commit(); conn.close()

def db_registrar_venta_final(mesa_id, detalle, total, metodo):
    conn = sqlite3.connect("pos_restaurante.db")
    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.cursor().execute("INSERT INTO ventas (mesa_id, detalle, total, fecha, metodo_pago, cerrada) VALUES (?, ?, ?, ?, ?, 0)", 
                           (mesa_id, detalle, total, ahora, metodo))
    conn.commit(); conn.close()

def db_ejecutar_cierre_caja():
    conn = sqlite3.connect("pos_restaurante.db")
    conn.cursor().execute("UPDATE ventas SET cerrada = 1 WHERE cerrada = 0")
    conn.commit(); conn.close()

def db_obtener_ventas_activas():
    conn = sqlite3.connect("pos_restaurante.db")
    datos = conn.cursor().execute("SELECT mesa_id, detalle, total, fecha, metodo_pago FROM ventas WHERE cerrada = 0").fetchall()
    conn.close(); return datos

def db_cargar_estado_inicial():
    conn = sqlite3.connect("pos_restaurante.db")
    cursor = conn.cursor()
    cursor.execute("SELECT mesa_id, nombre, precio, cantidad, destino, enviado FROM items_activos")
    filas = cursor.fetchall(); conn.close()
    datos = {i: [] for i in range(1, 21)}
    for f in filas: datos[f[0]].append({"n": f[1], "p": f[2], "q": f[3], "d": f[4], "enviado": bool(f[5])})
    return datos

# --- FUNCIÓN EXCEL ---
def generar_excel_cierre(ventas, total, efectivo, tarjeta):
    nombre_subcarpeta = "Reportes Cierre de Caja"
    if not os.path.exists(nombre_subcarpeta): os.makedirs(nombre_subcarpeta)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Corte de Caja"
    ws["A1"] = f"CORTE DE CAJA - TABLET {TABLET_ID}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A4"], ws["B4"] = "TOTAL GENERAL:", total
    ws["A5"], ws["B5"] = "EFECTIVO:", efectivo
    ws["A6"], ws["B6"] = "TARJETA:", tarjeta
    headers = ["Mesa", "Hora", "Método", "Total", "Detalle"]
    for col, text in enumerate(headers, 1): ws.cell(row=8, column=col, value=text).font = Font(bold=True)
    for row_idx, v in enumerate(ventas, 9):
        ws.cell(row=row_idx, column=1, value=f"Mesa {v[0]}")
        ws.cell(row=row_idx, column=2, value=v[3])
        ws.cell(row=row_idx, column=3, value=v[4])
        ws.cell(row=row_idx, column=4, value=v[2])
        ws.cell(row=row_idx, column=5, value=v[1])
    nombre_archivo = f"Corte_T{TABLET_ID}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    ruta_completa = os.path.join(nombre_subcarpeta, nombre_archivo)
    wb.save(ruta_completa); return ruta_completa

# --- APLICACIÓN PRINCIPAL ---
def main(page: ft.Page):
    init_db()
    page.title = "POS Restaurante Pro - SUNMI D3"
    page.theme_mode = "light"
    page.padding = 0

    cuentas = db_cargar_estado_inicial()
    estado = {"mesa": 0}

    # --- FUNCIONES DE LÓGICA ---
    def ocultar_todo():
        v_mesas.visible = v_pedido.visible = v_login.visible = False
        v_admin.visible = v_confirmacion.visible = v_ticket_final.visible = False
        v_confirm_cierre.visible = v_resumen_cierre.visible = v_pago_metodo.visible = False
        v_pago_finalizado.visible = v_gestion_menu.visible = False
        columna_botones_acciones.visible = True

    def salir_de_la_app(e):
        print("\n[!] Cerrando aplicación de forma segura...")
        import os
        os._exit(0)

    def ir_a_mesas(e):
        ocultar_todo()
        user_input.value = ""; pass_input.value = ""
        v_mesas.visible = True
        for c in grid_mesas.controls: c.bgcolor = "orange" if len(cuentas[c.data]) > 0 else "blue"
        page.update()

    def ir_a_login(e):
        ocultar_todo()
        user_input.value = ""; pass_input.value = ""
        v_login.visible = True
        page.update()

    def ir_a_pedido(e):
        ocultar_todo()
        estado["mesa"] = e.control.data
        txt_titulo_mesa.value = f"MESA #{estado['mesa']}"
        v_pedido.visible = True
        mostrar_mensaje_central("¡Bienvenido!\nSelecciona productos.", "blue")
        refrescar_ticket()
        page.update()

    def ir_a_admin(e):
        ocultar_todo()
        v_admin.visible = True
        actualizar_reporte_admin()
        page.update()

    def ir_a_gestion_menu(e):
        ocultar_todo()
        v_gestion_menu.visible = True
        refrescar_lista_gestion()
        page.update()

    def mostrar_mensaje_central(texto, color_texto):
        grid_prods.controls.clear()
        grid_prods.controls.append(ft.Column([ft.Container(height=100), ft.Row([ft.Text(texto, size=22, color=color_texto, weight="bold", text_align="center")], alignment=ft.MainAxisAlignment.CENTER)], horizontal_alignment=ft.CrossAxisAlignment.CENTER, width=700))
        page.update()

    # --- GESTIÓN DE PRODUCTOS (CON EDICIÓN DE PRECIO) ---
    def refrescar_lista_gestion():
        col_lista_prods.controls.clear()
        conn = sqlite3.connect("pos_restaurante.db")
        prods = conn.cursor().execute("SELECT id, nombre, precio, categoria FROM productos").fetchall()
        conn.close()
        for p in prods:
            txt_edit_precio = ft.TextField(value=str(p[2]), width=100, height=40, text_size=14)
            col_lista_prods.controls.append(
                ft.Row([
                    ft.Text(f"{p[1]} ({p[3]})", expand=True, size=16),
                    txt_edit_precio,
                    ft.TextButton("ACTUALIZAR", on_click=lambda e, id_p=p[0], tf=txt_edit_precio: actualizar_precio_db(id_p, tf.value)),
                    ft.TextButton("BORRAR", on_click=lambda e, idx=p[0]: eliminar_producto_menu(idx), style=ft.ButtonStyle(color="red"))
                ])
            )
        page.update()

    def actualizar_precio_db(id_p, nuevo_valor):
        try:
            nuevo_precio = float(nuevo_valor)
            conn = sqlite3.connect("pos_restaurante.db")
            conn.cursor().execute("UPDATE productos SET precio=? WHERE id=?", (nuevo_precio, id_p))
            conn.commit(); conn.close()
            refrescar_lista_gestion()
        except:
            pass

    def eliminar_producto_menu(idx):
        conn = sqlite3.connect("pos_restaurante.db")
        conn.cursor().execute("DELETE FROM productos WHERE id=?", (idx,))
        conn.commit(); conn.close()
        refrescar_lista_gestion()

    def añadir_nuevo_producto(e):
        if not txt_nom.value or not txt_pre.value: return
        conn = sqlite3.connect("pos_restaurante.db")
        conn.cursor().execute("INSERT INTO productos (nombre, precio, categoria, destino) VALUES (?,?,?,?)", 
                               (txt_nom.value, float(txt_pre.value), dd_cat.value, dd_dest.value))
        conn.commit(); conn.close()
        txt_nom.value = ""; txt_pre.value = ""
        refrescar_lista_gestion()

    # --- FILTRADO DE MENÚ DESDE DB ---
    def filtrar_menu_dinamico(categoria):
        grid_prods.controls.clear()
        conn = sqlite3.connect("pos_restaurante.db")
        prods = conn.cursor().execute("SELECT nombre, precio, destino FROM productos WHERE categoria=?", (categoria,)).fetchall()
        conn.close()
        btns = ft.GridView(runs_count=3, spacing=10, max_extent=150)
        for p in prods:
            btns.controls.append(ft.ElevatedButton(content=ft.Text(f"{p[0]}\n${p[1]}", text_align="center"),
                                 on_click=lambda e, n=p[0], pr=p[1], d=p[2]: agregar_item(n, pr, d), height=80))
        grid_prods.controls.append(btns); page.update()

    # --- LÓGICA PEDIDOS (LOGS DE COMANDA RESTAURADOS) ---
    def refrescar_ticket():
        col_ticket.controls.clear(); total = 0
        for item in cuentas[estado["mesa"]]:
            sub = item["p"] * item["q"]
            total += sub
            icono = ft.Text(" ✔ ", color="green") if item["enviado"] else ft.TextButton(content=ft.Text(" X ", color="red", weight="bold"), on_click=lambda e, n=item["n"]: quitar_item(n))
            col_ticket.controls.append(ft.Row([icono, ft.Text(f"{item['q']}x {item['n']}", expand=True), ft.Text(f"${sub}")]))
        txt_total.value = f"TOTAL: ${total}"; page.update()

    def quitar_item(nombre):
        m_id = estado["mesa"]
        for i, item in enumerate(cuentas[m_id]):
            if item["n"] == nombre and not item["enviado"]:
                db_remover_item_activo(m_id, nombre)
                if item["q"] > 1: item["q"] -= 1
                else: cuentas[m_id].pop(i)
                break
        refrescar_ticket()

    def agregar_item(nombre, precio, destino):
        m_id = estado["mesa"]; encontrado = False
        for item in cuentas[m_id]:
            if item["n"] == nombre and not item["enviado"]:
                item["q"] += 1; encontrado = True; db_guardar_item_activo(m_id, item); break
        if not encontrado:
            nuevo = {"n": nombre, "p": precio, "d": destino, "q": 1, "enviado": False}
            cuentas[m_id].append(nuevo); db_guardar_item_activo(m_id, nuevo)
        refrescar_ticket()

    def enviar_comanda(e):
        m_id = estado["mesa"]
        nuevos = [i for i in cuentas[m_id] if not i["enviado"]]
        if not nuevos: mostrar_mensaje_central("AVISO:\nNo hay productos nuevos.", "orange"); return
        
        # LOGS DE APERTURA RESTAURADOS
        if not any(i["enviado"] for i in cuentas[m_id]):
            print("\n" + "*"*50)
            print(f" [!] ESTADO: MESA {m_id} OCUPADA")
            print(f" [!] APERTURA: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print("*"*50)

        db_marcar_enviados(m_id)
        for i in cuentas[m_id]: i["enviado"] = True
        
        # LOGS DE COMANDA GRUPALES RESTAURADOS
        print(f"[*] ENVIANDO COMANDA MESA {m_id}")
        for dest in ["BARRA", "COCINA", "OTROS"]:
            items_batalla = [f"{i['q']}x {i['n']}" for i in nuevos if i["d"] == dest]
            if items_batalla: print(f"    -> {dest}: {', '.join(items_batalla)}")
            
        refrescar_ticket(); mostrar_mensaje_central("¡ORDEN ENVIADA!", "green")

    # --- ADVERTENCIAS Y PAGO (LOGS DE CIERRE RESTAURADOS) ---
    def validar_pago_antes_de_confirmar(e):
        mesa_actual = cuentas[estado["mesa"]]
        if not mesa_actual:
            mostrar_mensaje_central("ERROR:\nLa cuenta está vacía.", "red"); return
        if any(not i['enviado'] for i in mesa_actual):
            mostrar_mensaje_central("ADVERTENCIA:\nHay items sin enviar a comanda.", "red"); return
        setattr(v_confirmacion, 'visible', True); page.update()

    def finalizar_pago_total(metodo):
        m_id = estado["mesa"]
        items = cuentas[m_id]
        total = sum(i['p'] * i['q'] for i in items)
        detalle_texto = "\n".join([f"• {i['q']}x {i['n']}" for i in items])
        
        # LOGS DE CIERRE RESTAURADOS
        print("\n" + "!"*45)
        print(f" REPORTE CIERRE MESA {m_id} | TOTAL: ${total}")
        print(f" MÉTODO: {metodo.upper()}")
        print(f" FECHA: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print("!"*45 + "\n")

        db_registrar_venta_final(m_id, detalle_texto, total, metodo)
        db_limpiar_mesa(m_id); cuentas[m_id] = []
        ocultar_todo(); txt_mensaje_despedida.value = f"¡PAGO REGISTRADO!\nMétodo: {metodo.upper()}"; v_pago_finalizado.visible = True; page.update()

    def ejecutar_cierre_final(e):
        ventas = db_obtener_ventas_activas()
        total = sum(v[2] for v in ventas); efe = sum(v[2] for v in ventas if v[4].lower() == "efectivo"); tar = sum(v[2] for v in ventas if v[4].lower() == "tarjeta")
        archivo = generar_excel_cierre(ventas, total, efe, tar)
        db_ejecutar_cierre_caja()
        v_confirm_cierre.visible = False; txt_resumen_cierre_total.value = f"INGRESO TOTAL: ${total}"; txt_resumen_efectivo.value = f"EFECTIVO: ${efe}"; txt_resumen_tarjeta.value = f"TARJETA: ${tar}"
        txt_resumen_cierre_fecha.value = f"FECHA Y HORA: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"; v_resumen_cierre.visible = True; page.update()

    def actualizar_reporte_admin():
        ventas = db_obtener_ventas_activas()
        col_reportes_dia.controls.clear()
        for v in ventas:
            # DETALLES RESTAURADOS (PRODUCTOS Y HORA)
            col_reportes_dia.controls.append(
                ft.Container(
                    content=ft.Column([
                        ft.Row([
                            ft.Text(f"MESA {v[0]}", weight="bold", size=18), 
                            ft.Text(f"PAGO: {v[4].upper()}", color="blue", weight="bold")
                        ], alignment="spaceBetween"),
                        ft.Text("Productos:", size=14, weight="bold"),
                        ft.Text(v[1], color="grey", italic=True), # Detalles de productos
                        ft.Row([
                            ft.Text(f"Total: ${v[2]}", color="green", weight="bold", size=16), 
                            ft.Text(f"Hora: {v[3]}", size=12, color="grey") # Fecha y Hora
                        ], alignment="spaceBetween")
                    ]), padding=15, border=ft.border.all(1, "grey"), border_radius=10, margin=ft.margin.only(bottom=10)
                )
            )
        txt_ingreso_total_dia.value = f"TOTAL EN CAJA: ${sum(x[2] for x in ventas)}"; page.update()

    # --- UI COMPONENTS ---
    user_input, pass_input = ft.TextField(label="Usuario", width=350), ft.TextField(label="Contraseña", password=True, width=350)
    v_login = ft.Container(content=ft.Row([ft.Column([ft.Text("ACCESO ADMIN", size=40, weight="bold"), user_input, pass_input, ft.ElevatedButton("ENTRAR", bgcolor="blue", color="white", width=350, height=50, on_click=lambda _: ir_a_admin(None) if user_input.value=="admin" and pass_input.value=="1234" else None), ft.TextButton("VOLVER", on_click=ir_a_mesas)], alignment=ft.MainAxisAlignment.CENTER, horizontal_alignment=ft.CrossAxisAlignment.CENTER)], alignment=ft.MainAxisAlignment.CENTER), visible=False, expand=True, bgcolor="white")

    col_reportes_dia, txt_ingreso_total_dia = ft.Column(scroll="always", expand=True), ft.Text("", size=25, weight="bold", color="green")
    v_admin = ft.Container(content=ft.Column([ft.Row([ft.Text("REPORTE DIARIO", size=30, weight="bold"), ft.Row([ft.ElevatedButton("PRODUCTOS", bgcolor="blue", color="white", on_click=ir_a_gestion_menu), ft.ElevatedButton("SALIR APP", bgcolor="red", color="white", on_click=salir_de_la_app), ft.ElevatedButton("CIERRE", bgcolor="orange", color="white", on_click=lambda _: [setattr(v_confirm_cierre, 'visible', True), page.update()]), ft.TextButton("SALIR", on_click=ir_a_mesas)])], alignment="spaceBetween"), ft.Divider(), col_reportes_dia, ft.Divider(), ft.Row([txt_ingreso_total_dia], alignment="center")]), visible=False, expand=True, padding=30, bgcolor="white")

    txt_nom, txt_pre = ft.TextField(label="Nombre", width=250), ft.TextField(label="Precio", width=150)
    dd_cat = ft.Dropdown(label="Categoría", width=200, options=[ft.dropdown.Option("BEBIDAS"), ft.dropdown.Option("COMIDA"), ft.dropdown.Option("POSTRES")], value="BEBIDAS")
    dd_dest = ft.Dropdown(label="Destino", width=200, options=[ft.dropdown.Option("BARRA"), ft.dropdown.Option("COCINA")], value="BARRA")
    col_lista_prods = ft.Column(scroll="always", expand=True)
    v_gestion_menu = ft.Container(content=ft.Column([ft.Row([ft.Text("GESTIONAR PRODUCTOS", size=30, weight="bold"), ft.ElevatedButton("VOLVER", on_click=ir_a_admin)]), ft.Row([txt_nom, txt_pre, dd_cat, dd_dest, ft.ElevatedButton("AÑADIR", on_click=añadir_nuevo_producto, bgcolor="green", color="white")]), ft.Divider(), col_lista_prods]), visible=False, expand=True, padding=30, bgcolor="white")

    v_confirm_cierre = ft.Container(content=ft.Row([ft.Column([ft.Text("¡ADVERTENCIA!", size=30, weight="bold", color="red"), ft.Text("Se resetearán los ingresos."), ft.Row([ft.ElevatedButton("SÍ", bgcolor="green", color="white", on_click=ejecutar_cierre_final, width=150), ft.ElevatedButton("NO", bgcolor="red", color="white", on_click=lambda _: [setattr(v_confirm_cierre, 'visible', False), page.update()], width=150)], alignment="center")], alignment="center", horizontal_alignment="center")], alignment="center"), visible=False, expand=True, bgcolor="rgba(255,255,255,0.95)")
    v_resumen_cierre = ft.Container(content=ft.Row([ft.Column([ft.Text("RESUMEN CIERRE", size=30, weight="bold"), txt_resumen_cierre_total := ft.Text("", size=30, color="green", weight="bold"), txt_resumen_efectivo := ft.Text(""), txt_resumen_tarjeta := ft.Text(""), txt_resumen_cierre_fecha := ft.Text("", size=18, weight="bold"), txt_archivo_creado := ft.Text("", size=12, italic=True), ft.ElevatedButton("CERRAR", on_click=ir_a_admin)], alignment="center", horizontal_alignment="center")], alignment="center"), visible=False, expand=True, bgcolor="white")

    grid_mesas = ft.GridView(expand=True, runs_count=5, spacing=15)
    for i in range(1, 21): grid_mesas.controls.append(ft.Container(content=ft.Text(f"{i}", color="white", weight="bold"), bgcolor="blue", border_radius=10, padding=20, on_click=ir_a_pedido, data=i))
    v_mesas = ft.Container(content=ft.Column([ft.Row([ft.Text("SALÓN", size=30, weight="bold", expand=True), ft.TextButton("ADMIN", on_click=ir_a_login)]), grid_mesas]), expand=True, padding=20, bgcolor="white")

    columna_botones_acciones = ft.Column([
        ft.ElevatedButton("COMANDA", bgcolor="orange", color="white", height=60, on_click=enviar_comanda, width=400),
        ft.ElevatedButton("PAGAR CUENTA", bgcolor="green", color="white", height=60, on_click=validar_pago_antes_de_confirmar, width=400)
    ])

    txt_titulo_mesa, col_ticket, txt_total, grid_prods = ft.Text("", size=25, weight="bold"), ft.Column(scroll="always", expand=True), ft.Text("TOTAL: $0", size=35, weight="bold", color="green"), ft.Column(expand=True)
    v_pedido = ft.Container(content=ft.Row([ft.Column([ft.TextButton("<- VOLVER", on_click=ir_a_mesas), ft.Row([ft.ElevatedButton("BEBIDAS", on_click=lambda _: filtrar_menu_dinamico("BEBIDAS")), ft.ElevatedButton("COMIDA", on_click=lambda _: filtrar_menu_dinamico("COMIDA")), ft.ElevatedButton("POSTRES", on_click=lambda _: filtrar_menu_dinamico("POSTRES"))]), grid_prods], expand=3), ft.Container(content=ft.Column([txt_titulo_mesa, ft.Divider(), col_ticket, ft.Divider(), txt_total, columna_botones_acciones]), expand=2, bgcolor="#F5F5F5", padding=20, border_radius=15)]), expand=True, visible=False, bgcolor="white")

    v_confirmacion = ft.Container(content=ft.Row([ft.Column([ft.Text("¿CONFIRMAR PAGO?", size=25, weight="bold"), ft.Row([ft.ElevatedButton("SÍ, PAGAR", bgcolor="green", color="white", on_click=lambda _: [col_resumen_final.controls.clear(), [col_resumen_final.controls.append(ft.Text(f"{i['q']}x {i['n']} ... ${i['p']*i['q']}")) for i in cuentas[estado['mesa']]], setattr(v_ticket_final, 'visible', True), setattr(v_confirmacion, 'visible', False), setattr(columna_botones_acciones, 'visible', False), page.update()], width=180, height=60), ft.ElevatedButton("NO", bgcolor="red", color="white", on_click=lambda _: [setattr(v_confirmacion, 'visible', False), page.update()], width=180, height=60)], alignment="center")], alignment="center", horizontal_alignment="center")], alignment="center"), visible=False, expand=True, bgcolor="rgba(255,255,255,0.9)")
    
    col_resumen_final = ft.Column(scroll="always", expand=True)
    v_ticket_final = ft.Container(content=ft.Column([ft.Text("TICKET", size=30, weight="bold"), col_resumen_final, ft.ElevatedButton("FINALIZAR", bgcolor="blue", color="white", width=400, height=80, on_click=lambda _: [ocultar_todo(), setattr(v_pago_metodo, 'visible', True), page.update()])], horizontal_alignment="center"), bgcolor="white", visible=False, expand=True, padding=50)

    v_pago_metodo = ft.Container(content=ft.Row([ft.Column([ft.Text("MÉTODO DE PAGO", size=30, weight="bold"), ft.ElevatedButton("EFECTIVO", bgcolor="green", color="white", width=400, height=70, on_click=lambda _: finalizar_pago_total("Efectivo")), ft.ElevatedButton("TARJETA", bgcolor="blue", color="white", width=400, height=70, on_click=lambda _: finalizar_pago_total("Tarjeta"))], alignment="center", horizontal_alignment="center")], alignment="center"), visible=False, expand=True, bgcolor="white")
    v_pago_finalizado = ft.Container(content=ft.Row([ft.Column([ft.Text("GRACIAS", size=40, weight="bold"), txt_mensaje_despedida := ft.Text("", size=22, text_align="center"), ft.ElevatedButton("CERRAR", bgcolor="blue", color="white", width=350, height=70, on_click=ir_a_mesas)], alignment="center", horizontal_alignment="center")], alignment="center"), visible=False, expand=True, bgcolor="white")

    page.add(ft.Stack([v_mesas, v_pedido, v_login, v_admin, v_confirmacion, v_ticket_final, v_confirm_cierre, v_resumen_cierre, v_pago_metodo, v_pago_finalizado, v_gestion_menu], expand=True))
    ir_a_mesas(None)

if __name__ == "__main__":
    try: ft.app(target=main)
    except Exception as e:
        print("\n[!] Aplicación terminada.")